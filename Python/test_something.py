import re
import timeit
from datetime import date

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook("./Excel/input.xlsx")
wb['ARV']['A1'].value = 'fuck'


wb.save("./Excel/temp.xlsx")