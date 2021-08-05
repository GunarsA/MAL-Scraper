import re
import timeit
from datetime import date

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook("./Excel/Input.xlsx")
ws = wb['ARV']
ws.auto_filter.add_sort_condition("A2:B52")

wb.save("./Excel/temp.xlsx")