import re
import timeit
from datetime import date
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# Wraper function to calculate runtime of a function (function is able to
# return value if needed)
def calculate_runtime(func):
    def inner_function(*args, **kwargs):
        begin = timeit.default_timer()
        return_value = func(*args, **kwargs)
        end = timeit.default_timer()
        print("Runtime of '" + func.__name__ + "' function : " 
            + str(end - begin) + "\n")
        return return_value
    return inner_function


# Scrapes data from 6 MyAnimeList top rankings and stores it in different
# worksheets. Function also adds "Excel functions" that calculate order of
# animanga and stores it in seperate sheet.
@calculate_runtime
def _scrape_worksheet(value_ws, order_ws):

    def _request_specific_website(ws):

        def _get_specific_url(ws_title):
            URL_DICTIONARY = {
                'ARV': 'https://myanimelist.net/topanime.php',
                'AMV': 'https://myanimelist.net/topanime.php?type=bypopularity',
                'AFV': 'https://myanimelist.net/topanime.php?type=favorite',
                'MRV': 'https://myanimelist.net/topmanga.php',
                'MMV': 'https://myanimelist.net/topmanga.php?type=bypopularity',
                'MFV': 'https://myanimelist.net/topmanga.php?type=favorite'
            }
            return URL_DICTIONARY.get(ws_title)
        
        return requests.get(_get_specific_url(ws.title)).text

    def _find_specific_animanga_data(ws, soup):
        if ws.title == 'ARV':
            title = soup.find('h3', class_
                = 'hoverinfo_trigger fl-l fs14 fw-b anime_ranking_h3').text
            core = soup.find('td', class_
                = 'score ac fs14').text.replace('\n','').replace(' ','')
        if ws.title == 'AMV':
            title = soup.find('h3', class_
                = 'hoverinfo_trigger fl-l fs14 fw-b anime_ranking_h3').text
            temp = soup.find(string = re.compile('members'))
            core = temp.replace(' ','').replace(',','')\
                .replace('members','').replace('\n','')
        if ws.title == 'AFV':
            title = soup.find('h3', class_
                = 'hoverinfo_trigger fl-l fs14 fw-b anime_ranking_h3').text
            temp = soup.find(string = re.compile('favorites'))
            core = temp.replace(' ','').replace(',','')\
                .replace('favorites','').replace('\n','')
        if ws.title == 'MRV':
            title = soup.find('h3', class_ = 'manga_h3').text
            core = soup.find('td', class_
                = 'score ac fs14').text.replace('\n','')
        if ws.title == 'MMV':
            title = soup.find('h3', class_ = 'manga_h3').text
            temp = soup.find(string = re.compile('members'))
            core = temp.replace(' ','').replace(',','')\
                .replace('members','').replace('\n','')
        if ws.title == 'MFV':
            title = soup.find('h3', class_ = 'manga_h3').text
            temp = soup.find(string = re.compile('favorites'))
            core = temp.replace(' ','').replace(',','')\
                .replace('favorites','').replace('\n','')
        return (title, core)

    def _add_data_to_worksheets(v_ws, o_ws, title, info, col_idx):

        def _get_worksheet_row_count(ws):
            row_count = 1
            while ws['A' + str(row_count)].value:
                    row_count += 1
            return row_count
        
        data_has_changed_for_animanga = False

        for i in range(2, _get_worksheet_row_count(v_ws)):
            if v_ws['B' + str(i)].value == title:
                v_ws[col_idx + str(i)] = info

                o_ws[col_idx + str(i)] = ('=COUNTIF(' + v_ws.title + '!$'
                    + col_idx + '$2:$' + col_idx + '$100,">"&' + v_ws.title
                    + '!' + col_idx + str(i) + ')+1')

                if v_ws[chr(ord(col_idx) - 2) + str(i)].value:
                    v_ws[chr(ord(col_idx) - 1) + str(i)].value = (info 
                        - v_ws[chr(ord(col_idx) - 2) + str(i)].value)
                    if info != v_ws[chr(ord(col_idx) - 2) + str(i)].value\
                        and (v_ws.title == 'ARV' or v_ws.title == 'MRV'):
                        print(title + " data changed: "
                            + str(v_ws[chr(ord(col_idx) - 2) + str(i)].value)
                            + " -> " + str(info))
                        data_has_changed_for_animanga = True

                    o_ws[chr(ord(col_idx) - 1) + str(i)].value = ('=' 
                        + (chr(ord(col_idx) - 2) + str(i)) + ' - '
                        + (col_idx + str(i)))
                else:
                    print(title + " data changed: NULL -> " + str(info))
                    data_has_changed_for_animanga = True
                    
                break

        else:
            print('+ New animanga added to ' + v_ws.title + ': ' + title + ' | '
                + str(info))
            ws_row_count = _get_worksheet_row_count(v_ws)

            v_ws['A' + str(ws_row_count)] = '#' + str(ws_row_count - 1)
            v_ws['B' + str(ws_row_count)] = title
            v_ws[col_idx + str(ws_row_count)] = info

            o_ws['A' + str(ws_row_count)] = '#' + str(ws_row_count - 1)
            o_ws['B' + str(ws_row_count)] = title
            o_ws[col_idx + str(ws_row_count)] = ('=COUNTIF(' + v_ws.title
                + '!$' + col_idx + '$2:$' + col_idx + '$100,">"&' + v_ws.title
                + '!' + col_idx + str(ws_row_count) + ')+1')
            
            data_has_changed_for_animanga = True

        return data_has_changed_for_animanga

    COLLUM_INDEX = get_column_letter(value_ws.max_column + 2)
    HTML_TEXT = _request_specific_website(value_ws)

    SOUP = BeautifulSoup(HTML_TEXT, 'lxml')
    RANKING_LIST = SOUP.find_all('tr', class_ = 'ranking-list')

    data_has_changed_for_ws = False

    for animanga_soup_data in RANKING_LIST:
        TEMP = _find_specific_animanga_data(value_ws, animanga_soup_data)
        if _add_data_to_worksheets(value_ws, order_ws, TEMP[0], float(TEMP[1]),
            COLLUM_INDEX) == True:
            data_has_changed_for_ws = True

    value_ws.auto_filter.ref = "A1:" + COLLUM_INDEX + str(value_ws.max_row)
    value_ws[COLLUM_INDEX + '1'] = date.today()
    value_ws[chr(ord(COLLUM_INDEX) - 1) + '1'] = 'Change'

    order_ws.auto_filter.ref = "A1:" + COLLUM_INDEX + str(value_ws.max_row)
    order_ws[COLLUM_INDEX + '1'] = date.today()
    order_ws[chr(ord(COLLUM_INDEX) - 1) + '1'] = 'Change'

    if not data_has_changed_for_ws:
        print("No major changes in worksheet")


# Program adds aditional data to premade data worksheets
def main():
    workbook = load_workbook("./input.xlsx")

    WORKSHEET_TITLE_DICTIONARY = {
        'ARV': 'ARO',
        'AMV': 'AMO',
        'AFV': 'AFO',
        'MRV': 'MRO',
        'MMV': 'MMO',
        'MFV': 'MFO'
    }
    
    TEMP = date.today()
    TODAYS_DATE = TEMP.strftime("%Y.%m.%d")

    for main_worksheet_title in WORKSHEET_TITLE_DICTIONARY:
        _scrape_worksheet(workbook[main_worksheet_title],
            workbook[WORKSHEET_TITLE_DICTIONARY[main_worksheet_title]])

    START_TIME = timeit.default_timer()
    workbook.save("./Excel/" + TODAYS_DATE + ".xlsx")
    END_TIME = timeit.default_timer()
    print("Saved as '" + TODAYS_DATE + ".xlsx" + "' in "
    + str(END_TIME - START_TIME))


if __name__ == '__main__':
    main()